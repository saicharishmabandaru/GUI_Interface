import tkinter as tk
import openpyxl
from array import array
import pandas as pd
import subprocess
import os


constants_entries = {}
constants_entries=dict()
observed_data_entries = {}
testarr = []
testarr1 = []
testarr2 = []

# Function to read data from Excel and populate the dictionaries and defaults
def read_data_from_excel():
    try:
        workbook = openpyxl.load_workbook("C:/Users/SAI CHARISHMA/Downloads/data_new.xlsx")
        constants_sheet = workbook["Constants"]
        observed_data_sheet = workbook["Observed Data"]
        Albedo_sheet=workbook["Albedo"]

        for row in constants_sheet.iter_rows(values_only=True):
            testarr.append(row)

        for row in observed_data_sheet.iter_rows(values_only=True):
            testarr1.append(row)
        
        for row in Albedo_sheet.iter_rows(values_only=True):
            testarr2.append(row) 

    except FileNotFoundError:
        print("Data Excel file not found. Creating a new one.")
    except Exception as e:
        print("An error occurred while reading data from Excel:", str(e))

# Call the function to populate the entry fields with Excel data
read_data_from_excel()


def write_observed_data_to_excel(updated_observed_data):
    try:
        excel_file_path = 'C:/Users/SAI CHARISHMA/Downloads/data_new.xlsx'
        sheet_name = 'Observed Data'
        target_column = 'B'  # Update this if the target column is different

        # Load the existing workbook
        existing_workbook1 = openpyxl.load_workbook(excel_file_path)

        # Select the sheet
        sheet1 = existing_workbook1[sheet_name]

        # Write the updated constants to the specified column in the Excel sheet
        for index, value in enumerate(updated_observed_data, start=2):
            sheet1[target_column + str(index)] = float(value)

        # Save the workbook
        existing_workbook1.save(excel_file_path)

        print("Observed data saved to Excel file successfully.")
    except Exception as e:
        print(f"An error occurred while saving constants: {e}")

def write_constants_to_excel(updated_constants):
    try:
        excel_file_path = 'C:/Users/SAI CHARISHMA/Downloads/data_new.xlsx'
        sheet_name = 'Constants'
        target_column = 'B'  # Update this if the target column is different

        # Load the existing workbook
        existing_workbook = openpyxl.load_workbook(excel_file_path)

        # Select the sheet
        sheet = existing_workbook[sheet_name]

        # Write the updated constants to the specified column in the Excel sheet
        for index, value in enumerate(updated_constants, start=2):
            sheet[target_column + str(index)] = float(value)


        # Save the workbook
        existing_workbook.save(excel_file_path)

        print("Constants saved to Excel file successfully.")
    except Exception as e:
        print(f"An error occurred while saving constants: {e}")

def write_data_to_excel(updated_albedo_value):
    try:
        workbook = openpyxl.load_workbook("C:/Users/SAI CHARISHMA/Downloads/data_new.xlsx")
        albedo_sheet = workbook["Albedo"]
        albedo_sheet['B2'] = updated_albedo_value
        workbook.save("C:/Users/SAI CHARISHMA/Downloads/data_new.xlsx")
        print("albedo value saved to Excel file successfully.")
    except Exception as e:
        print(f"An error occurred: {e}")

# Function to handle button click event
def save_to_excel():
    try:
        # Get the updated albedo value from the GUI (assuming albedo_entry is a widget)
        updated_albedo_value = float(albedo_entries["A1"].get())  # Assuming entry is for numeric values

        # Call the function to save data to Excel
        write_data_to_excel(updated_albedo_value)

        # Get the updated constants from the GUI
        updated_constants = []

        for param in constants_parameters:
            updated_constants.append(constants_entries[param].get())

        # Call the function to save constants to Excel
        write_constants_to_excel(updated_constants)

        # Get the updated constants from the GUI
        updated_observed_data = []

        for param in observed_data_parameters:
            updated_observed_data.append(observed_data_entries[param].get())

        # Call the function to save constants to Excel
        write_observed_data_to_excel(updated_observed_data)

        print("Data saved to Excel file successfully.")
    except ValueError:
        print("Invalid input. Make sure numeric values are entered.")
    except Exception as e:
        print(f"An error occurred: {e}")
        
# Create the main window
root = tk.Tk()
root.title("Glacier Energy Mass Balance Model")

# Add the heading at the top of the window
heading_label = tk.Label(root, text="Inputs for the Hypsometric Glacier Energy Mass Balance Model (HGEMBM)", font=("Times New Roman", 20, "bold"))
heading_label.grid(row=0, column=0, columnspan=2, pady=(10, 20))

# Create a side heading for CONSTANTS
constants_heading_label = tk.Label(root, text="CONSTANTS", font=("Helvetica", 14, "bold"))
constants_heading_label.grid(row=1, column=0, padx=20, pady=(10, 5), sticky='w')

# CONSTANTS section
constants_frame = tk.Frame(root)
constants_frame.grid(row=2, column=0, padx=20, pady=5, sticky='w')

# Create labels and entry fields for each constant parameter with 6 cm distance from the left side
constants_parameters = [
    "Thickness interval of ice (bi,m)", "Thickness interval of snow (bs,m)", "Bulk coefficient (c,N/A)",
    "Specific heat of air (ca,J/kg-K)", "Specific heat of ice (ci,J/kg-K)",
    "Latent heat of evaporation of ice (le,J/kg)", "Latent heat of fusion of ice (lf,J/kg)","Latent heat of vaporization(lv,J/kg)",
    "Extinction coefficient of ice (mui,/m)", "Extinction coefficient of snow (mus,/m)",
    "Stefan-boltzmann constant (B,W/m^2-K^4)", "Density of ice (di,kgm^-3)","Density of water (dw,kgm^-3)"
]

i=1
for index, param in enumerate(constants_parameters):
    label = tk.Label(constants_frame, text=param)
    label.grid(row=index, column=0, sticky='w', padx=120, pady=5)  
    entry = tk.Entry(constants_frame)
    entry.grid(row=index, column=1, padx=20, pady=5, sticky='w') 
    entry.insert(0,testarr[i][1]) 
    constants_entries[param] = entry
    i+=1

# Create a side heading for OBSERVED DATA
observed_data_heading_label = tk.Label(root, text="OBSERVED DATA", font=("Helvetica", 14, "bold"))
observed_data_heading_label.grid(row=1, column=1, padx=20, pady=(10, 5), sticky='w')

# OBSERVED DATA section
observed_data_frame = tk.Frame(root)
observed_data_frame.grid(row=2, column=1, padx=20, pady=5, sticky='w')

# Create labels and entry fields for each observed data parameter with 6 cm distance from the left side
observed_data_parameters = [
    " Air temperature (Ta,degree Celcius)", " Relative Humidity (rh)", 
    " Observed downward shortwave radiation (Rs,Wm^-2)", "Wind speed (U,ms^-1)", 
    "Daily precipitation (Pp,mm.w.e.d^-1)", "Emissivity of snow/ice surface (e)", 
    " Downward longwave radiation (Rl)", "Density of air (da,kgm^-3)", 
    "Saturated specific humidity (q)", "Assumption- specific humidity (G_q)", 
    "Density of snow (ds,kgm^-3)"
]

i=1
observed_data_entries = {}
for index, param in enumerate(observed_data_parameters):
    label = tk.Label(observed_data_frame, text=param)
    label.grid(row=index, column=0, sticky='w', padx=100, pady=5)  
    entry = tk.Entry(observed_data_frame)
    entry.grid(row=index, column=1, padx=20, pady=5, sticky='w')  
    # print(testarr[i])
    entry.insert(0,testarr1[i][1])
    observed_data_entries[param] = entry
    i+=1

# Add a heading for INPUT on the left side
input_heading_label = tk.Label(root, text="MAIN INPUT", font=("Helvetica", 16, "bold"))
input_heading_label.grid(row=3, column=0, pady=(20, 10))

# Create a Frame for Albedo label and entry below INPUT
albedo_frame = tk.Frame(root)
albedo_frame.grid(row=4, column=0, pady=(10, 20))

Albedo_parameters = [
    "A1"
]

# Create an Entry box for "Albedo"
i=1
albedo_entries = {}
for index, param in enumerate(Albedo_parameters):
   label = tk.Label(albedo_frame, text="Albedo")
   label.grid(row=0, column=0, padx=10, sticky='w')
   entry = tk.Entry(albedo_frame)
   entry.grid(row=0, column=1, padx=10, sticky='w')
   entry.insert(0,testarr2[i][1])
   #print(testarr2[i][1])
   albedo_entries[param] = entry


# Add a heading for OUTPUT on the right side
output_heading_label = tk.Label(root, text="OUTPUT", font=("Helvetica", 16, "bold"))
output_heading_label.grid(row=3, column=1, pady=(20, 10))


# Add a heading for OUTPUT on the right side
output_heading_label = tk.Label(root, text="OUTPUT", font=("Helvetica", 16, "bold"))
output_heading_label.grid(row=3, column=1, pady=(20, 10))

# Create a Frame for Runoff label and entry below OUTPUT
runoff_frame = tk.Frame(root)
runoff_frame.grid(row=4, column=1, pady=(10, 20))

# Create a Label for "Runoff"
runoff_label = tk.Label(runoff_frame, text="Runoff(mm.w.e)")
runoff_label.grid(row=0, column=0, padx=10, sticky='w')

# Create an Entry box for "Runoff"
runoff_entry = tk.Entry(runoff_frame)
runoff_entry.grid(row=0, column=1, padx=10, sticky='w')


# Function to handle Reset button click event
def on_reset():
    try:
        # Load the original data from the new Excel file
        workbook = openpyxl.load_workbook("C:/Users/SAI CHARISHMA/Downloads/data_reset.xlsx")
        constants_sheet = workbook["Constants"]
        observed_data_sheet = workbook["Observed Data"]
        Albedo_sheet = workbook["Albedo"]

        testarr = [row for row in constants_sheet.iter_rows(values_only=True)]
        testarr1 = [row for row in observed_data_sheet.iter_rows(values_only=True)]
        testarr2 = [row for row in Albedo_sheet.iter_rows(values_only=True)]

        # Update the constants entry fields with the original data
        i = 1
        for index, param in enumerate(constants_parameters):
            constants_entries[param].delete(0, tk.END)
            constants_entries[param].insert(0, testarr[i][1])  # Change index here
            i += 1

        # Update the observed data entry fields with the original data
        i = 1
        for index, param in enumerate(observed_data_parameters):
            observed_data_entries[param].delete(0, tk.END)
            observed_data_entries[param].insert(0, testarr1[i][1])  # Change index here
            i += 1

        # Update the albedo entry field with the original data
        albedo_entries["A1"].delete(0, tk.END)
        albedo_entries["A1"].insert(0, testarr2[1][1])  # Change index here

        print("Data reset successfully.")
    except Exception as e:
        print(f"An error occurred during reset: {e}")

# Function to handle Save to Excel button click event
def on_save():
    # Implement saving to Excel functionality here if needed
    pass


def on_run():
        
        # Specify the path to your MATLAB script
        matlab_script_path = r"C:\Users\SAI CHARISHMA\Downloads\recent.m"

        # Specify the path to your MATLAB executable
        matlab_executable_path = r'C:\Program Files\MATLAB\R2023b\bin\matlab.exe'

        # Normalize the path
        normalized_matlab_script_path = os.path.normpath(matlab_script_path)

        matlab_command = [
        matlab_executable_path,
        '-batch',
        f'run(\'{normalized_matlab_script_path}\')', 
    ]


        # Capture the MATLAB console output
        result = subprocess.run(matlab_command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

        print("Data saved to Excel file and MATLAB model run successfully.")

        try:
            # Get the updated albedo value from the GUI
            updated_albedo_str = albedo_entries["A1"].get()

            # Check if the input is not empty and contains only numeric characters or a single decimal point
            if updated_albedo_str and updated_albedo_str.replace('.', '', 1).isdigit():
                updated_albedo_value = float(updated_albedo_str)

                # Define the file path with the correct escape sequence for backslashes
                file_path = r"C:\Users\SAI CHARISHMA\Downloads\Runoff_Albedo.csv"

                # Load the CSV file containing runoff-albedo data without header
                runoff_albedo_data = pd.read_csv(file_path, header=None)

                # Find the row where the albedo matches the updated_albedo_value
                matching_row = runoff_albedo_data[runoff_albedo_data.iloc[:, 0] == updated_albedo_value]

                if not matching_row.empty:
                    # Get the corresponding runoff value
                    runoff_value = matching_row.iloc[0, 1]

                    # Update the runoff entry field with the retrieved runoff value
                    runoff_entry.delete(0, tk.END)
                    runoff_entry.insert(0, runoff_value)
                    print("Runoff value retrieved successfully.")
                else:
                    print("No matching runoff value found for the given albedo.")
            else:
                print("Invalid input. Make sure numeric values are entered.")
        except FileNotFoundError:
            print("Runoff-albedo CSV file not found at the specified path.")
        except Exception as e:
            print(f"An unexpected error occurred: {e}")


# Create a Frame for buttons below the Runoff section
button_frame = tk.Frame(root)
button_frame.grid(row=5, column=0, columnspan=2, pady=20)

# Create buttons for Reset, Save to Excel, and Run inside the button frame
reset_button = tk.Button(button_frame, text="Reset", command=on_reset)
reset_button.grid(row=0, column=0, padx=10)

save_button = tk.Button(button_frame, text="Save to Excel", command=save_to_excel)
save_button.grid(row=0, column=1, padx=10)

run_button = tk.Button(button_frame, text="Run",command=on_run)
run_button.grid(row=0, column=2, padx=10)

# Start the main loop
root.mainloop()